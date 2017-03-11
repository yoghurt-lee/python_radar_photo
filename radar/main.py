#coding:utf-8

from openpyxl import Workbook,load_workbook
import random,os
import numpy as np
import matplotlib.pyplot as plt

class Excel(object):
    def __init__(self):
        self.path = 'data/'
        if not os.path.exists(self.path):
            os.makedirs(self.path)

    def create_excel(self,sheet_name_list,data):
        wb  = Workbook()
        ws=[]
        for i in range(len(sheet_name_list)):
            ws.append(wb.create_sheet(title=sheet_name_list[i].decode()))
        for i in range(len(sheet_name_list)):
            ws[i].append(['序号','体力','精神','智力','力量','幸运度'])
            count = 1
            for j in range(5):
                ws[i].append([count,data[j][0],data[j][1],data[j][2],data[j][3],data[j][4]])
                count+=1
        save_path = self.path+'need'
        for i in range(len(sheet_name_list)):
            save_path+=('-'+sheet_name_list[i].decode())
        save_path+='.xlsx'
        if os.path.exists(save_path):
            print u"该标签文件已存在..."
            return
        wb.save(save_path)

    def read_from_excel(self,path):
        wb = load_workbook(filename=path)
        ws = wb.get_sheet_by_name('sheet1')
        info_id = []
        info_title = []
        info_data = []
        for i in range(2,7):
            info_title.append(ws.cell(row=1,column=i).value)

        for i in range(2, 7):  ## 遍历第2行到6 行
            id = ws.cell(row=i, column=1).value  ## 遍历第2行到6 行，第1列
            info_id.append(id)

        for row in range(2, len(info_id) + 2):  ## 遍历第2行到7行
            row_empty = []  ##建立一个空数组作为临时储存地，每次换行就被清空
            for i in range(2, 7):  ## 遍历第2行到32行，第2到9列
                data_excel = ws.cell(row=row, column=i).value
                if data_excel == None:
                    pass
                else:
                    row_empty.append(data_excel)  ##将单元格信息储存进去
            info_data.append(row_empty)
        return [info_title,info_id,info_data]

    def draw(self,labels,info_id,info_data):
        # numpy 中文资料: http://old.sebug.net/paper/books/scipydoc/numpy_intro.html
        #设置雷达各个顶点的名称
        labels = np.array(labels)
        #数据个数
        data_len = len(labels)
        for i in range(0,len(info_id)):
            data = np.array(info_data[i])
            # numpy.linspace(start, stop, num=50, endpoint=True, retstep=False, dtype=None) 生成等比数列
            # start是采样的起始点
            #stop是采样的终点
            #num是采样的点个数
            #endpoint关键字指定是否包括终值，缺省设置是包括终值.(绘制这个图的话可以认为是最后一个点不会和第一个点结合)
            #retstep会改变计算的输出，输出一个元组，而元组的两个元素分别是需要生成的数列和数列的步进差值。
            angles = np.linspace(0, 2*np.pi, data_len, endpoint=False)
            #numpy.concatenate((a1,a2,...), axis=0):能够一次完成多个数组的拼接。其中a1,a2,...是数组类型的参数(这里就是第一个数据和最后一个数据)
            data = np.concatenate((data, [data[0]])) # 闭合
            angles = np.concatenate((angles, [angles[0]])) # 闭合
            # 调用figure创建一个绘图对象，并且使它成为当前的绘图对象
            fig = plt.figure()
            # add_subplot(numRows, numCols, plotNum)#将图片做成numRows行,numCols列,图像画在从左到右从上到下的第plotNum块
            ax = fig.add_subplot(1,1,1, polar=True)# polar参数！！polar为true用来绘制极坐标图
            ax.plot(angles, data, 'bo-', linewidth=2)# 画线
            ax.fill(angles, data, facecolor='red', alpha=0.25)# 填充
            ax.set_thetagrids(angles * 180/np.pi, labels, fontproperties="SimHei")
            ax.set_title("属性:" + str(info_id[i]), va='bottom', fontproperties="SimHei")
            ax.set_rlim(3.8,5)# 设置雷达图的范围(从中心到外围)
            ax.grid(True)
            plt.savefig(self.path + str(info_id[i]) + ".png", dpi=120)

    def reduce_pictrue(self,infile,outfile):
        from PIL import Image
        ##更改图片大小
        im = Image.open(infile)
        (x, y) = im.size
        #长宽分别缩小为原来的1/2
        x1 = x/2
        y2 = y/2
        out = im.resize((x1, y2), Image.ANTIALIAS)
        out.save(outfile,'png',quality = 95)

if __name__ == '__main__':
    excel = Excel()
    data = []
    for i in range(5):
        temp = []
        for j in range(5):
            temp.append(round(random.uniform(4,5),2))
        data.append(temp)
    excel.create_excel(['sheet1'],data)
    info_first,info_id,info_data = excel.read_from_excel('data/need-sheet1.xlsx')
    excel.draw(info_first,info_id,info_data)
    excel.reduce_pictrue('data/1.png','data/1.1.png')
